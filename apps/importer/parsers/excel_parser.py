"""Excel file parser (.xlsx / .xls) using openpyxl."""

import logging

from .base import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parse Excel files into structured row data."""

    def parse_file(self, file_path_or_buffer, sheet_name=None) -> ParseResult:
        import openpyxl

        errors = []
        try:
            wb = openpyxl.load_workbook(
                file_path_or_buffer, read_only=True, data_only=True
            )
        except Exception as e:
            return ParseResult(errors=[f"Failed to open Excel file: {e}"])

        # Pick sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows_data = []
        headers = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                # First row = headers
                headers = [self._clean_header(h) for h in row]
                continue

            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_dict = {}
            for j, value in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = self._clean_value(value)

            rows_data.append(row_dict)

        wb.close()

        logger.info("Excel parsed: %d headers, %d rows", len(headers), len(rows_data))
        return ParseResult(
            headers=headers,
            rows=rows_data,
            errors=errors,
            metadata={"sheet": ws.title, "format": "excel"},
        )
